#!/usr/bin/env python3
"""
Template Analyzer for Subsidy Application Forms

このスクリプトはExcelテンプレートを詳細分析し、
以下の情報をJSON形式で出力します:
- シート構成
- セル結合パターン
- Named Range
- 動的表の検出
- 数式セル
"""

import openpyxl
from openpyxl.utils import get_column_letter
import json
import sys
from pathlib import Path


class TemplateAnalyzer:
    def __init__(self, template_path):
        self.template_path = template_path
        self.wb = openpyxl.load_workbook(template_path, data_only=False)
        self.analysis = {
            "file_name": Path(template_path).name,
            "sheets": [],
            "sheet_details": {}
        }
    
    def analyze(self):
        """完全な分析を実行"""
        print(f"📊 テンプレート分析開始: {self.template_path}")
        print("=" * 60)
        
        # シート一覧
        self.analysis["sheets"] = self.wb.sheetnames
        
        # 各シートを分析
        for sheet_name in self.wb.sheetnames:
            print(f"\n[{sheet_name}]")
            self.analysis["sheet_details"][sheet_name] = self._analyze_sheet(sheet_name)
        
        # Named Range分析
        self.analysis["named_ranges"] = self._analyze_named_ranges()
        
        return self.analysis
    
    def _analyze_sheet(self, sheet_name):
        """個別シートの詳細分析"""
        ws = self.wb[sheet_name]
        
        sheet_info = {
            "dimensions": {
                "max_row": ws.max_row,
                "max_column": ws.max_column
            },
            "merged_cells": self._analyze_merged_cells(ws),
            "formula_cells": self._find_formula_cells(ws),
            "tables": self._detect_tables(ws),
            "data_validation": self._find_data_validation(ws)
        }
        
        return sheet_info
    
    def _analyze_merged_cells(self, ws):
        """セル結合の詳細分析"""
        merged_details = []
        
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left = f"{get_column_letter(min_col)}{min_row}"
            
            # 結合セルの現在値を取得
            current_value = ws[top_left].value
            
            merged_details.append({
                "range": str(merged_range),
                "write_to": top_left,
                "rows": max_row - min_row + 1,
                "cols": max_col - min_col + 1,
                "current_value": str(current_value) if current_value else None,
                "note": "⚠️ 常に左上セル(write_to)に書き込むこと"
            })
        
        print(f"  セル結合: {len(merged_details)}箇所")
        return merged_details
    
    def _find_formula_cells(self, ws):
        """数式セルを検出(上書き禁止)"""
        formula_cells = []
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # formula
                    formula_cells.append({
                        "cell": cell.coordinate,
                        "formula": str(cell.value),
                        "note": "🚫 数式セル - 上書き禁止"
                    })
        
        print(f"  数式セル: {len(formula_cells)}個")
        return formula_cells
    
    def _detect_tables(self, ws):
        """動的表(ヘッダー + データ行)を検出"""
        tables = []
        current_table = None
        
        for row_idx in range(1, ws.max_row + 1):
            row = list(ws[row_idx])
            
            # 行に値が含まれているかチェック
            non_empty_cells = [cell for cell in row if cell.value is not None]
            
            if len(non_empty_cells) >= 3:  # 3列以上に値がある
                # ヘッダーらしき行を検出
                header_values = [str(cell.value) for cell in non_empty_cells]
                
                # 「項目」「金額」「数量」などのキーワードがあればヘッダーと判定
                keywords = ['項目', '金額', '数量', '単価', '名称', '内容', '備考']
                is_header = any(kw in val for kw in keywords for val in header_values)
                
                if is_header and current_table is None:
                    current_table = {
                        "header_row": row_idx,
                        "header_values": header_values,
                        "data_start_row": row_idx + 1,
                        "columns": [cell.column_letter for cell in non_empty_cells]
                    }
            
            # 空白行でテーブル終了
            elif len(non_empty_cells) == 0 and current_table:
                current_table["data_end_row"] = row_idx - 1
                tables.append(current_table)
                current_table = None
        
        # 最後のテーブル処理
        if current_table:
            current_table["data_end_row"] = ws.max_row
            tables.append(current_table)
        
        print(f"  動的表検出: {len(tables)}個")
        for i, table in enumerate(tables):
            print(f"    表{i+1}: {table['header_row']}行目～ (列: {', '.join(table['columns'])})")
        
        return tables
    
    def _find_data_validation(self, ws):
        """データ検証(プルダウン等)を検出"""
        validations = []
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            for dv in ws.data_validations.dataValidation:
                validations.append({
                    "cells": str(dv.sqref),
                    "type": dv.type,
                    "formula": dv.formula1 if hasattr(dv, 'formula1') else None
                })
        
        if validations:
            print(f"  データ検証: {len(validations)}箇所")
        
        return validations
    
    def _analyze_named_ranges(self):
        """Named Range(名前付き範囲)を分析"""
        named_ranges = []

        for defn in self._iter_defined_names():
            name = getattr(defn, "name", None)
            if not name:
                continue

            destinations = list(defn.destinations)
            if destinations:
                sheet, coord = destinations[0]
                
                # セルの値を取得
                try:
                    ws = self.wb[sheet]
                    cell_value = ws[coord].value
                except:
                    cell_value = None
                
                named_ranges.append({
                    "name": name,
                    "sheet": sheet,
                    "cell": coord,
                    "current_value": str(cell_value) if cell_value else None,
                    "purpose_hint": self._guess_purpose(name)
                })
        
        print(f"\nNamed Range: {len(named_ranges)}個")
        for nr in named_ranges:
            print(f"  {nr['name']} → {nr['sheet']}!{nr['cell']}")
        
        return named_ranges

    def _iter_defined_names(self):
        """openpyxl のバージョン差異を吸収して DefinedName を列挙する"""
        defined_names = self.wb.defined_names

        if hasattr(defined_names, "definedName"):
            return list(defined_names.definedName)

        if hasattr(defined_names, "values"):
            return list(defined_names.values())

        return list(defined_names)
    
    def _guess_purpose(self, name):
        """Named Rangeの用途を推測"""
        name_lower = name.lower()
        
        if 'company' in name_lower or '会社' in name_lower:
            if 'name' in name_lower or '名' in name_lower:
                return "会社名"
            elif 'address' in name_lower or '住所' in name_lower:
                return "会社住所"
            elif 'tel' in name_lower or '電話' in name_lower:
                return "電話番号"
        
        if 'project' in name_lower or '事業' in name_lower:
            if 'name' in name_lower or 'title' in name_lower:
                return "事業名"
        
        if 'date' in name_lower or '日付' in name_lower:
            return "日付"
        
        if 'amount' in name_lower or '金額' in name_lower:
            return "金額"
        
        return "不明"
    
    def save_to_file(self, output_path):
        """分析結果をJSONファイルに保存"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.analysis, f, indent=2, ensure_ascii=False)
        
        print(f"\n✅ 分析完了: {output_path}")
    
    def print_summary(self):
        """分析結果のサマリーを表示"""
        print("\n" + "=" * 60)
        print("📋 分析サマリー")
        print("=" * 60)
        
        total_merged = sum(
            len(details["merged_cells"]) 
            for details in self.analysis["sheet_details"].values()
        )
        total_formulas = sum(
            len(details["formula_cells"]) 
            for details in self.analysis["sheet_details"].values()
        )
        total_tables = sum(
            len(details["tables"]) 
            for details in self.analysis["sheet_details"].values()
        )
        
        print(f"シート数: {len(self.analysis['sheets'])}")
        print(f"セル結合: {total_merged}箇所")
        print(f"数式セル: {total_formulas}個")
        print(f"動的表: {total_tables}個")
        print(f"Named Range: {len(self.analysis['named_ranges'])}個")
        
        print("\n⚠️ 重要な注意点:")
        print("  1. セル結合がある場合、必ず左上セル(write_to)に書き込む")
        print("  2. 数式セルは絶対に上書きしない")
        print("  3. 動的表は行数に応じて動的に埋め込む")


def main():
    if len(sys.argv) < 2:
        print("使い方: python analyze_template.py <テンプレートファイル.xlsx> [出力先.json]")
        sys.exit(1)
    
    template_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else "template_analysis.json"
    
    analyzer = TemplateAnalyzer(template_path)
    analyzer.analyze()
    analyzer.print_summary()
    analyzer.save_to_file(output_path)


if __name__ == "__main__":
    main()
