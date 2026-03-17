#!/usr/bin/env python3
"""
Template Filler

マッピング定義に従って、抽出データをExcelテンプレートに埋め込みます。
セル結合、Named Range、動的表に完全対応。
"""

import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import sys
from pathlib import Path
from datetime import datetime


class TemplateFiller:
    def __init__(self, template_path, mapping_def_path, extracted_data_path):
        """
        Args:
            template_path: Excelテンプレートのパス
            mapping_def_path: マッピング定義JSONのパス
            extracted_data_path: 抽出データJSONのパス
        """
        self.template_path = template_path
        
        # テンプレートを読み込み
        self.wb = openpyxl.load_workbook(template_path)
        
        # マッピング定義を読み込み
        with open(mapping_def_path, 'r', encoding='utf-8') as f:
            self.mapping_def = json.load(f)
        
        # 抽出データを読み込み
        with open(extracted_data_path, 'r', encoding='utf-8') as f:
            self.extracted_data = json.load(f)
        
        self.fill_log = []

    EXAMPLE_TEXT_PATTERNS = [
        r"記入例",
        r"入力例",
        r"例[：:]",
        r"サンプル",
        r"ダミー",
        r"ここに",
        r"入力してください",
        r"記載してください",
        r"株式会社〇〇",
        r"株式会社XX",
        r"山田太郎",
        r"○○",
        r"XXX",
        r"ＡＡＡ",
    ]
    
    def fill(self):
        """データ埋め込みを実行"""
        print("📝 データ埋め込み開始...")
        print("=" * 60)
        
        mappings = self.mapping_def.get("mappings", {})
        
        for sheet_name, sheet_mappings in mappings.items():
            if sheet_name not in self.wb.sheetnames:
                print(f"⚠️ シート '{sheet_name}' が見つかりません")
                continue
            
            print(f"\n[{sheet_name}]")
            ws = self.wb[sheet_name]
            
            # 1. シンプルなセル埋め込み
            if "simple_cells" in sheet_mappings:
                self._fill_simple_cells(ws, sheet_mappings["simple_cells"])
            
            # 2. Named Range埋め込み
            if "named_ranges" in sheet_mappings:
                self._fill_named_ranges(sheet_mappings["named_ranges"])
            
            # 3. 動的表埋め込み
            if "dynamic_tables" in sheet_mappings:
                self._fill_dynamic_tables(ws, sheet_mappings["dynamic_tables"])
        
        print("\n" + "=" * 60)
        print("✅ 埋め込み完了")
    
    def _get_value_from_path(self, data_path):
        """
        data_path (例: "company_info.name") から値を取得
        
        Returns:
            値(str, int, float等) または None
        """
        keys = data_path.split(".")
        value = self.extracted_data
        
        for key in keys:
            if isinstance(value, dict):
                value = value.get(key)
            elif isinstance(value, list):
                if not key.isdigit():
                    return None
                index = int(key)
                if index < 0 or index >= len(value):
                    return None
                value = value[index]
            else:
                return None
        
        return value
    
    def _fill_simple_cells(self, ws, simple_cells):
        """シンプルなセルへの値埋め込み"""
        print("  📌 単純セル埋め込み:")
        
        for mapping in simple_cells:
            cell_coord = mapping["cell"]
            data_path = mapping["data_path"]
            description = mapping.get("description", "")
            
            value = self._get_value_from_path(data_path)
            
            if value is None:
                print(f"    ⚠️ {cell_coord} ({description}): データなし")
                continue
            
            # セル結合チェック
            target_cell = self._get_write_target(ws, cell_coord)

            self._clear_example_text_in_cell(ws, target_cell)

            # 値を設定
            ws[target_cell] = value
            
            # フォーマット適用
            self._apply_format(ws[target_cell], mapping.get("format"))
            
            print(f"    ✓ {target_cell} ← {value} ({description})")
            self.fill_log.append({
                "sheet": ws.title,
                "cell": target_cell,
                "value": value,
                "description": description
            })
    
    def _get_write_target(self, ws, cell_coord):
        """
        セル結合を考慮した書き込み対象セルを取得
        
        セル結合されている場合は左上セルを返す
        """
        for merged_range in ws.merged_cells.ranges:
            if cell_coord in merged_range:
                # 結合範囲の左上セルを取得
                min_col, min_row, _, _ = merged_range.bounds
                return f"{get_column_letter(min_col)}{min_row}"
        
        return cell_coord
    
    def _apply_format(self, cell, format_spec):
        """セルにフォーマットを適用"""
        if not format_spec:
            return
        
        if format_spec == "date":
            cell.number_format = 'YYYY/MM/DD'
        elif format_spec == "number":
            cell.number_format = '#,##0'
        elif format_spec == "currency":
            cell.number_format = '¥#,##0'
        elif format_spec == "percentage":
            cell.number_format = '0.0%'
    
    def _fill_named_ranges(self, named_ranges):
        """Named Rangeへの値埋め込み"""
        print("  📌 Named Range埋め込み:")
        
        for mapping in named_ranges:
            name = mapping["name"]
            data_path = mapping["data_path"]
            description = mapping.get("description", "")
            
            value = self._get_value_from_path(data_path)
            
            if value is None:
                print(f"    ⚠️ {name} ({description}): データなし")
                continue
            
            # Named Rangeの参照先を取得
            if name not in self.wb.defined_names:
                print(f"    ⚠️ {name}: Named Rangeが見つかりません")
                continue
            
            defn = self.wb.defined_names[name]
            destinations = list(defn.destinations)
            
            if not destinations:
                print(f"    ⚠️ {name}: 参照先が不明")
                continue
            
            sheet_name, coord = destinations[0]
            ws = self.wb[sheet_name]
            
            # セル結合チェック
            target_cell = self._get_write_target(ws, coord)

            self._clear_example_text_in_cell(ws, target_cell)

            ws[target_cell] = value
            
            # フォーマット適用
            self._apply_format(ws[target_cell], mapping.get("format"))
            
            print(f"    ✓ {name} ({target_cell}) ← {value} ({description})")
            self.fill_log.append({
                "sheet": sheet_name,
                "named_range": name,
                "cell": target_cell,
                "value": value,
                "description": description
            })
    
    def _fill_dynamic_tables(self, ws, dynamic_tables):
        """動的表への埋め込み(最も複雑)"""
        print("  📌 動的表埋め込み:")
        
        for table_config in dynamic_tables:
            table_id = table_config.get("table_id", "unnamed")
            print(f"    テーブル: {table_id}")
            
            data_path = table_config["data_path"]
            table_data = self._get_value_from_path(data_path)
            
            if not isinstance(table_data, list):
                print(f"      ⚠️ {data_path} はリストではありません")
                continue
            
            if len(table_data) == 0:
                print(f"      ⚠️ データが空です")
                continue
            
            data_start_row = table_config["data_start_row"]
            row_step = table_config.get("row_step", 1)
            columns = table_config["columns"]
            
            # 既存データ行をクリア(オプション)
            if table_config.get("clear_existing", False):
                self._clear_table_rows(ws, table_config)
            
            # データ行を埋め込み
            for row_offset, item in enumerate(table_data):
                current_row = data_start_row + (row_offset * row_step)
                
                for col_config in columns:
                    col_letter = col_config["col"]
                    data_field = col_config["data_field"]
                    
                    value = item.get(data_field, "")
                    
                    cell_coord = f"{col_letter}{current_row}"
                    target_cell = self._get_write_target(ws, cell_coord)

                    self._clear_example_text_in_cell(ws, target_cell)

                    ws[target_cell] = value
                    
                    # フォーマット適用
                    self._apply_format(ws[target_cell], col_config.get("format"))
                
                print(f"      ✓ 行{current_row}: {item.get(columns[0]['data_field'], '')}")
            
            # 合計行の処理
            if "auto_sum" in table_config:
                self._add_sum_row(ws, table_config, len(table_data))
    
    def _clear_table_rows(self, ws, table_config):
        """既存のテーブルデータ行をクリア"""
        data_start_row = table_config["data_start_row"]
        data_end_row = table_config.get("data_end_row", data_start_row + 100)
        columns = table_config["columns"]
        
        for row in range(data_start_row, data_end_row + 1):
            for col_config in columns:
                col_letter = col_config["col"]
                ws[f"{col_letter}{row}"] = None

    def _looks_like_example_text(self, value):
        """入力欄に残った例文・ダミーテキストらしさを判定する"""
        if not isinstance(value, str):
            return False

        text = value.strip()
        if not text:
            return False

        return any(re.search(pattern, text, re.IGNORECASE) for pattern in self.EXAMPLE_TEXT_PATTERNS)

    def _clear_example_text_in_cell(self, ws, cell_coord):
        """書き込み前に、入力欄に残っている例文を削除する"""
        cell = ws[cell_coord]
        if self._looks_like_example_text(cell.value):
            print(f"    🧹 {cell_coord} の例文を削除")
            cell.value = None
    
    def _add_sum_row(self, ws, table_config, data_count):
        """合計行を追加"""
        auto_sum = table_config["auto_sum"]
        data_start_row = table_config["data_start_row"]
        
        sum_row = data_start_row + data_count + auto_sum.get("row_offset", 0)
        sum_col = auto_sum["col"]
        sum_cell = f"{sum_col}{sum_row}"
        
        # 数式を生成
        end_row = data_start_row + data_count - 1
        formula_template = auto_sum.get("formula_template", "=SUM({col}{start}:{col}{end})")
        formula = formula_template.format(
            col=sum_col,
            start=data_start_row,
            end=end_row
        )
        
        ws[sum_cell] = formula
        
        # フォーマット適用
        if "format" in auto_sum:
            self._apply_format(ws[sum_cell], auto_sum["format"])
        
        print(f"      ✓ 合計 ({sum_cell}): {formula}")
    
    def save(self, output_path):
        """完成ファイルを保存"""
        self.wb.save(output_path)
        print(f"\n💾 保存完了: {output_path}")
    
    def verify(self):
        """埋め込み結果を検証"""
        print("\n" + "=" * 60)
        print("🔍 検証中...")
        print("=" * 60)
        
        issues = []
        
        # 主要セルが埋まっているか確認
        for log_entry in self.fill_log:
            if "cell" in log_entry:
                cell = log_entry["cell"]
                # シート名が含まれていない場合は最初のシートを使用
                if "!" in cell:
                    sheet_name, cell_coord = cell.split("!")
                    ws = self.wb[sheet_name]
                elif "sheet" in log_entry:
                    ws = self.wb[log_entry["sheet"]]
                    cell_coord = cell
                else:
                    ws = self.wb.active
                    cell_coord = cell
                
                if ws[cell_coord].value is None:
                    issues.append(f"⚠️ {cell} が空です")
                elif self._looks_like_example_text(ws[cell_coord].value):
                    issues.append(f"⚠️ {cell} に例文が残っています")
        
        if issues:
            print("検証で問題が見つかりました:")
            for issue in issues:
                print(f"  {issue}")
            return False
        else:
            print("✅ 検証OK - すべてのセルに値が埋まっています")
            return True


def main():
    if len(sys.argv) < 4:
        print("使い方: python fill_template.py <テンプレート.xlsx> <マッピング定義.json> <抽出データ.json> [出力先.xlsx]")
        print("\n例:")
        print("  python fill_template.py template.xlsx mapping.json extracted_data.json output.xlsx")
        sys.exit(1)
    
    template_path = sys.argv[1]
    mapping_path = sys.argv[2]
    data_path = sys.argv[3]
    output_path = sys.argv[4] if len(sys.argv) > 4 else "completed_application.xlsx"
    
    # ファイル存在チェック
    for path, name in [(template_path, "テンプレート"), (mapping_path, "マッピング定義"), (data_path, "抽出データ")]:
        if not Path(path).exists():
            print(f"❌ エラー: {name}ファイルが見つかりません: {path}")
            sys.exit(1)
    
    # 埋め込み実行
    filler = TemplateFiller(template_path, mapping_path, data_path)
    filler.fill()
    
    # 検証
    is_valid = filler.verify()
    
    # 保存
    filler.save(output_path)
    
    if not is_valid:
        print("\n⚠️ 検証で問題が見つかりましたが、ファイルは保存されました。")
        print("   手動で確認してください。")
    else:
        print("\n🎉 完成! 申請書が正常に生成されました。")


if __name__ == "__main__":
    main()
