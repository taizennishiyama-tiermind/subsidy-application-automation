#!/usr/bin/env python3
"""財務分析の簡易 Excel 出力。"""

import argparse
import json
from openpyxl import Workbook
from openpyxl.styles import Font


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-data", required=True)
    parser.add_argument("--ratios", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    with open(args.source_data, "r", encoding="utf-8") as f:
        source = json.load(f)
    with open(args.ratios, "r", encoding="utf-8") as f:
        ratios = json.load(f)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "3期比較"
    ws1["A1"] = "3期比較財務諸表"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.append(["年度", "売上高", "営業利益", "経常利益", "当期純利益", "付加価値額"])
    for period, ratio in zip(source.get("periods", []), ratios.get("ratios", [])):
        ws1.append([
            period.get("year"),
            period.get("sales"),
            period.get("operating_profit"),
            period.get("ordinary_profit"),
            period.get("net_profit"),
            ratio.get("value_added"),
        ])

    ws2 = wb.create_sheet("指標")
    ws2["A1"] = "主要財務指標"
    ws2["A1"].font = Font(bold=True, size=14)
    headers = ["年度", "営業利益率", "ROA", "ROE", "自己資本比率", "流動比率", "債務償還年数"]
    ws2.append(headers)
    for ratio in ratios.get("ratios", []):
        ws2.append([
            ratio.get("year"),
            ratio.get("operating_profit_margin"),
            ratio.get("roa"),
            ratio.get("roe"),
            ratio.get("equity_ratio"),
            ratio.get("current_ratio"),
            ratio.get("debt_repayment_years"),
        ])

    ws3 = wb.create_sheet("評価メモ")
    ws3["A1"] = "申請書へ転用する評価"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3.append(["年度", "要点"])
    for ratio in ratios.get("ratios", []):
        note = f"自己資本比率={ratio.get('evaluations', {}).get('equity_ratio', '')} / 営業利益率={ratio.get('evaluations', {}).get('operating_profit_margin', '')}"
        ws3.append([ratio.get("year"), note])

    wb.save(args.output)
    print(f"OK: wrote {args.output}")


if __name__ == "__main__":
    main()
