"""
ものづくり補助金 申請書 完成版Excel生成スクリプト
デモ案件_山田製作所 用
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# データ読み込み
# ─────────────────────────────────────────────
BASE = os.path.dirname(__file__)
DATA_FILE = os.path.join(BASE, "application_data.json")
OUTPUT_DIR = os.path.join(BASE, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "ものづくり補助金_申請書_山田製作所_完成版.xlsx")

with open(DATA_FILE, encoding="utf-8") as f:
    data = json.load(f)

c = data["cover"]
bp = data["business_plan"]
nt = data["numerical_targets"]
wi = data["wage_increase"]
fp = data["funding_plan"]

# ─────────────────────────────────────────────
# スタイル定義
# ─────────────────────────────────────────────
wb = Workbook()

# 色定義
HEADER_BG  = "1F4E79"   # 濃紺
SUBHDR_BG  = "2E75B6"   # 中紺
LABEL_BG   = "D6E4F0"   # 薄青
YELLOW_BG  = "FFF2CC"   # 黄色（強調）
GREEN_BG   = "E2EFDA"   # 薄緑（数値目標）
WHITE      = "FFFFFF"

def make_font(bold=False, size=11, color="000000", name="游ゴシック"):
    return Font(bold=bold, size=size, color=color, name=name)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(thick=False):
    s = Side(style="medium" if thick else "thin")
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(ws, row, col, value, colspan=1, bg=HEADER_BG, size=12):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, size=size, color=WHITE, name="游ゴシック")
    cell.fill = make_fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border(thick=True)
    if colspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + colspan - 1
        )

def label_cell(ws, row, col, value, bg=LABEL_BG):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = make_font(bold=True, size=10)
    cell.fill = make_fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = make_border()
    return cell

def value_cell(ws, row, col, value, colspan=1, fmt=None, bg=WHITE, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = make_font(bold=bold, size=10)
    cell.fill = make_fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = make_border()
    if fmt:
        cell.number_format = fmt
    if colspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + colspan - 1
        )
    return cell

# ─────────────────────────────────────────────
# シート1: 表紙
# ─────────────────────────────────────────────
ws1 = wb.active
ws1.title = "表紙"
ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 50

# タイトル
ws1.merge_cells("A1:B1")
t = ws1["A1"]
t.value = "ものづくり補助金 申請書"
t.font = Font(bold=True, size=16, color=WHITE, name="游ゴシック")
t.fill = make_fill(HEADER_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:B2")
t2 = ws1["A2"]
t2.value = "【 省力化・デジタル枠 】"
t2.font = Font(bold=True, size=12, color=WHITE, name="游ゴシック")
t2.fill = make_fill(SUBHDR_BG)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 24

rows = [
    ("申請日",          c["application_date"]),
    ("会社名",          c["company_name"]),
    ("代表者名",        c["representative_name"]),
    ("所在地",          c["address"]),
    ("電話番号",        c["phone"]),
    ("資本金（万円）",   f'{c["capital"]:,}万円'),
    ("従業員数（名）",   f'{c["employee_count"]}名'),
    ("業種",            c["industry"]),
    ("事業計画名",      c["project_title"]),
    ("補助事業区分",    c["subsidy_type"]),
    ("補助希望額（万円）", f'{c["subsidy_amount"]:,}万円'),
    ("事業実施期間",    c["implementation_period"]),
]

for i, (lbl, val) in enumerate(rows, start=4):
    ws1.row_dimensions[i].height = 22
    label_cell(ws1, i, 1, lbl)
    vc = value_cell(ws1, i, 2, val)
    if lbl == "事業計画名":
        vc.font = Font(bold=True, size=10, name="游ゴシック")
        vc.fill = make_fill(YELLOW_BG)

# ─────────────────────────────────────────────
# シート2: 事業計画書
# ─────────────────────────────────────────────
ws2 = wb.create_sheet("事業計画書")
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 80

header_style(ws2, 1, 1, "事業計画書", colspan=2, size=14)
ws2.row_dimensions[1].height = 30

sections = [
    ("1-1", "現在の事業の状況・強み・弱みや市場動向等について",  bp["business_current_situation"],  "800字以内"),
    ("1-2", "革新的サービス・試作品・生産プロセス等の開発の内容", bp["innovation_content"],           "1,200字以内"),
    ("1-3", "補助事業の実施体制",                               bp["implementation_structure"],    "600字以内"),
    ("1-4", "補助事業の実施スケジュール",                       bp["implementation_schedule"],     "600字以内"),
    ("2-1", "本事業の成果の事業化に向けて取り組む内容",           bp["commercialization_plan"],      "800字以内"),
    ("2-2", "市場規模・成長性等について",                        bp["market_analysis"],             "600字以内"),
    ("2-3", "現在の自社の製品・サービス等と比較した優位性",        bp["competitive_advantage"],       "600字以内"),
]

row = 2
for sec_num, title, body, limit in sections:
    # セクションヘッダー
    header_style(ws2, row, 1, f"【{sec_num}】{title}（{limit}）", colspan=2, bg=SUBHDR_BG, size=10)
    ws2.row_dimensions[row].height = 22
    row += 1

    # 本文
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws2.cell(row=row, column=1, value=body)
    cell.font = make_font(size=10)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = make_border()
    # 行高さ: 文字数に応じて調整（1行あたり約50文字、1行=15pt想定）
    estimated_lines = max(len(body) // 50 + 2, 8)
    ws2.row_dimensions[row].height = estimated_lines * 15
    row += 2

# 数値目標セクション
header_style(ws2, row, 1, "【数値目標】付加価値額・労働生産性", colspan=2, bg=HEADER_BG, size=11)
ws2.row_dimensions[row].height = 22
row += 1

tbl_headers = ["指標", "現状値", "3年後目標値", "伸び率"]
for ci, h in enumerate(tbl_headers, start=1):
    header_style(ws2, row, ci, h, bg=SUBHDR_BG, size=10)
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 18
ws2.row_dimensions[row].height = 18
row += 1

num_rows = [
    ("付加価値額", f'{nt["added_value_current"]:,}万円/年', f'{nt["added_value_3years"]:,}万円/年', f'+{nt["added_value_growth_3y"]}%'),
    ("1人あたり付加価値額（労働生産性）", f'{nt["productivity_per_employee_current"]:,}万円/人', f'{nt["productivity_per_employee_3years"]:,}万円/人', "+41.3%"),
]
for lbl, cur, tgt, gro in num_rows:
    value_cell(ws2, row, 1, lbl, bg=LABEL_BG)
    value_cell(ws2, row, 2, cur, bg=GREEN_BG)
    value_cell(ws2, row, 3, tgt, bg=GREEN_BG, bold=True)
    value_cell(ws2, row, 4, gro, bg=YELLOW_BG, bold=True)
    ws2.row_dimensions[row].height = 20
    row += 1

# ─────────────────────────────────────────────
# シート3: 賃金引上げ計画
# ─────────────────────────────────────────────
ws3 = wb.create_sheet("賃金引上げ計画")
ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 22
ws3.column_dimensions["D"].width = 22

header_style(ws3, 1, 1, "賃金引上げ計画", colspan=4, size=13)
ws3.row_dimensions[1].height = 28

# ヘッダー行
for ci, h in enumerate(["項目", "現状", "引上げ後", "引上げ額"], start=1):
    header_style(ws3, 2, ci, h, bg=SUBHDR_BG, size=10)
ws3.row_dimensions[2].height = 18

# データ行
wage_rows = [
    ("事業場内最低賃金（時給）",
     f'{wi["min_wage_current"]:,}円',
     f'{wi["min_wage_after"]:,}円以上',
     f'+{wi["min_wage_after"] - wi["min_wage_current"]:,}円（+3.0%以上）'),
    ("対象従業員数", f'{wi["target_employees"]}名', "同左", "—"),
    ("実施時期", "—", wi["implementation_timing"], "—"),
]
for i, (lbl, cur, aft, diff) in enumerate(wage_rows, start=3):
    ws3.row_dimensions[i].height = 20
    label_cell(ws3, i, 1, lbl)
    value_cell(ws3, i, 2, cur)
    value_cell(ws3, i, 3, aft, bg=GREEN_BG)
    value_cell(ws3, i, 4, diff, bg=YELLOW_BG if i == 3 else WHITE, bold=(i == 3))

# 注記
ws3.merge_cells("A7:D7")
note = ws3["A7"]
note.value = "※ 補助事業期間内（2026年度）に全従業員（32名）の賃金を平均3%以上引き上げることをコミットします。"
note.font = Font(size=9, name="游ゴシック", color="555555")
note.alignment = Alignment(wrap_text=True)
ws3.row_dimensions[7].height = 30

# ─────────────────────────────────────────────
# シート4: 資金調達計画
# ─────────────────────────────────────────────
ws4 = wb.create_sheet("資金調達計画")
ws4.column_dimensions["A"].width = 52
ws4.column_dimensions["B"].width = 22
ws4.column_dimensions["C"].width = 18

header_style(ws4, 1, 1, "資金調達計画", colspan=3, size=13)
ws4.row_dimensions[1].height = 28

header_style(ws4, 2, 1, "補助対象経費明細", colspan=3, bg=SUBHDR_BG, size=10)
ws4.row_dimensions[2].height = 18

for ci, h in enumerate(["設備名・内容", "金額（税抜）", "備考"], start=1):
    header_style(ws4, 3, ci, h, bg=SUBHDR_BG, size=10)
ws4.row_dimensions[3].height = 18

equip_rows = [
    (fp["equipment_name_1"], f'{fp["equipment_cost_1"]:,}万円', "精密加工・省力化の中核設備"),
    (fp["equipment_name_2"], f'{fp["equipment_cost_2"]:,}万円', "AI検査・デジタル化"),
    (fp["equipment_name_3"], f'{fp["equipment_cost_3"]:,}万円', "MES・生産管理・設置費含む"),
]
for i, (name, cost, note) in enumerate(equip_rows, start=4):
    ws4.row_dimensions[i].height = 22
    value_cell(ws4, i, 1, name)
    value_cell(ws4, i, 2, cost, bg=GREEN_BG)
    value_cell(ws4, i, 3, note)

# 合計行
ws4.row_dimensions[7].height = 24
label_cell(ws4, 7, 1, "補助対象経費 合計")
value_cell(ws4, 7, 2, f'{fp["total_eligible_cost"]:,}万円', bg=YELLOW_BG, bold=True)
ws4.cell(row=7, column=3).value = ""

# 空行
ws4.row_dimensions[8].height = 10

# 補助・自己負担
header_style(ws4, 9, 1, "補助金・自己負担", colspan=3, bg=SUBHDR_BG, size=10)
ws4.row_dimensions[9].height = 18

funding_rows = [
    ("補助申請額（対象経費の2/3）",    f'{fp["subsidy_amount"]:,}万円',   "上限4,500万円以内"),
    ("自己負担額（対象経費の1/3）",    f'{fp["self_burden"]:,}万円',       ""),
    ("資金調達①：自己資金",           f'{fp["funding_self"]:,}万円',      "現預金から充当"),
    ("資金調達②：借入",               f'{fp["funding_loan"]:,}万円',      "借入なし"),
]
for i, (lbl, amt, note) in enumerate(funding_rows, start=10):
    ws4.row_dimensions[i].height = 20
    label_cell(ws4, i, 1, lbl)
    value_cell(ws4, i, 2, amt, bg=(YELLOW_BG if i == 10 else WHITE), bold=(i == 10))
    value_cell(ws4, i, 3, note)

# ─────────────────────────────────────────────
# 保存
# ─────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"✅ 完成版Excel を生成しました:\n   {OUTPUT_FILE}")
