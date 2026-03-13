"""テンプレート差し込み位置の解決と検証を行う共通ユーティリティ。

固定セルだけに依存せず、ラベル文言やシート分類を使って
「今回投入されたテンプレートでどこへ書くべきか」を毎回確認する。
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


REFERENCE_SHEET_KEYWORDS = (
    "記入例",
    "記入方法",
    "説明",
    "見本",
    "サンプル",
    "入力例",
    "注意",
    "参考",
)

DEFAULT_OFFSETS = (
    (0, 1),
    (0, 2),
    (0, 3),
    (1, 0),
    (1, 1),
    (1, 2),
    (1, 3),
)


class MappingResolutionError(Exception):
    """マッピング解決に失敗したときの例外。"""


@dataclass
class ResolvedField:
    sheet_name: str
    field_name: str
    cell_ref: str
    data_key: str
    required: bool
    resolver: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("\n", " ")
    text = re.sub(r"\s+", "", text)
    text = text.replace("（", "(").replace("）", ")")
    return text


def is_reference_sheet(sheet_name: str) -> bool:
    return any(keyword in sheet_name for keyword in REFERENCE_SHEET_KEYWORDS)


def resolve_merged_cell(ws: Worksheet, cell_ref: str):
    cell = ws[cell_ref]
    if not isinstance(cell, MergedCell):
        return cell
    for merge_range in ws.merged_cells.ranges:
        if cell.coordinate in merge_range:
            return ws.cell(row=merge_range.min_row, column=merge_range.min_col)
    return cell


def _merged_bounds_for_cell(ws: Worksheet, row: int, col: int) -> tuple[int, int, int, int]:
    for merge_range in ws.merged_cells.ranges:
        if (
            merge_range.min_row <= row <= merge_range.max_row
            and merge_range.min_col <= col <= merge_range.max_col
        ):
            return (
                merge_range.min_row,
                merge_range.min_col,
                merge_range.max_row,
                merge_range.max_col,
            )
    return row, col, row, col


def _looks_like_label(text: str, labels: list[str]) -> bool:
    normalized = normalize_text(text)
    if not normalized:
        return False
    for label in labels:
        label_norm = normalize_text(label)
        if label_norm and (label_norm in normalized or normalized in label_norm):
            return True
    return False


def _is_formula_cell(cell) -> bool:
    return getattr(cell, "data_type", None) == "f" or (
        isinstance(cell.value, str) and cell.value.startswith("=")
    )


def _find_anchor_matches(ws: Worksheet, labels: list[str]) -> list[tuple[int, Any]]:
    matches: list[tuple[int, Any]] = []
    normalized_labels = [normalize_text(label) for label in labels if normalize_text(label)]
    if not normalized_labels:
        return matches

    for row in ws.iter_rows():
        for cell in row:
            current = normalize_text(cell.value)
            if not current:
                continue
            best_score = 0
            for label in normalized_labels:
                if current == label:
                    best_score = max(best_score, 100)
                elif label in current:
                    best_score = max(best_score, 90)
                elif current in label:
                    best_score = max(best_score, 70)
            if best_score:
                matches.append((best_score, cell))
    matches.sort(key=lambda item: (-item[0], item[1].row, item[1].column))
    return matches


def _find_candidate_near_anchor(
    ws: Worksheet,
    anchor_cell,
    labels: list[str],
    preferred_offsets: list[list[int]] | None = None,
):
    offsets = preferred_offsets or DEFAULT_OFFSETS
    min_row, _, max_row, max_col = _merged_bounds_for_cell(ws, anchor_cell.row, anchor_cell.column)

    for row_offset, col_offset in offsets:
        candidate_row = min_row + row_offset
        candidate_col = max_col + col_offset
        if candidate_row < 1 or candidate_col < 1:
            continue
        if candidate_row > ws.max_row or candidate_col > ws.max_column:
            continue
        candidate = resolve_merged_cell(ws, f"{get_column_letter(candidate_col)}{candidate_row}")
        if candidate.coordinate == anchor_cell.coordinate:
            continue
        if _is_formula_cell(candidate):
            continue
        if _looks_like_label(candidate.value, labels):
            continue
        return candidate
    return None


def _normalize_entry(field_name: str, entry: Any) -> dict[str, Any]:
    if isinstance(entry, str):
        return {
            "data_key": entry,
            "targets": [],
            "required": False,
            "anchors": [],
            "preferred_offsets": None,
        }
    if "data_key" not in entry:
        raise MappingResolutionError(f"{field_name}: data_key がありません")
    return {
        "data_key": entry["data_key"],
        "targets": entry.get("targets", []),
        "required": bool(entry.get("required", False)),
        "anchors": entry.get("anchors", []),
        "preferred_offsets": entry.get("preferred_offsets"),
    }


def build_sheet_cell_map(workbook, mapping_profile: dict, data: dict) -> tuple[dict, dict]:
    sheet_cell_map: dict[str, dict[str, Any]] = {}
    report = {
        "resolved_fields": [],
        "reference_sheets_skipped": [],
        "missing_required_data": [],
        "unresolved_required_fields": [],
        "unresolved_optional_fields": [],
    }

    for sheet_name, fields in mapping_profile.items():
        if sheet_name not in workbook.sheetnames:
            report["unresolved_required_fields"].append(f"{sheet_name}: シート未存在")
            continue
        if is_reference_sheet(sheet_name):
            report["reference_sheets_skipped"].append(sheet_name)
            continue

        ws = workbook[sheet_name]
        resolved_cells: dict[str, Any] = {}

        for field_name, raw_entry in fields.items():
            entry = _normalize_entry(field_name, raw_entry)
            data_key = entry["data_key"]
            value = data.get(data_key)
            if value is None or value == "":
                if entry["required"]:
                    report["missing_required_data"].append(f"{sheet_name}.{field_name} -> {data_key}")
                continue

            chosen_cell = None
            resolver = None

            for target in entry["targets"]:
                candidate = resolve_merged_cell(ws, target)
                if _is_formula_cell(candidate):
                    continue
                chosen_cell = candidate.coordinate
                resolver = f"direct:{target}"
                break

            if chosen_cell is None and entry["anchors"]:
                for _, anchor_cell in _find_anchor_matches(ws, entry["anchors"]):
                    candidate = _find_candidate_near_anchor(
                        ws,
                        anchor_cell,
                        entry["anchors"],
                        preferred_offsets=entry["preferred_offsets"],
                    )
                    if candidate is not None:
                        chosen_cell = candidate.coordinate
                        resolver = f"anchor:{anchor_cell.coordinate}"
                        break

            if chosen_cell is None:
                bucket = (
                    "unresolved_required_fields" if entry["required"] else "unresolved_optional_fields"
                )
                report[bucket].append(f"{sheet_name}.{field_name} -> {data_key}")
                continue

            resolved_cells[chosen_cell] = value
            report["resolved_fields"].append(
                ResolvedField(
                    sheet_name=sheet_name,
                    field_name=field_name,
                    cell_ref=chosen_cell,
                    data_key=data_key,
                    required=entry["required"],
                    resolver=resolver or "unknown",
                ).__dict__
            )

        if resolved_cells:
            sheet_cell_map[sheet_name] = resolved_cells

    return sheet_cell_map, report


def validate_resolution_report(report: dict):
    errors: list[str] = []
    if report["missing_required_data"]:
        errors.append("必須データ不足: " + ", ".join(report["missing_required_data"]))
    if report["unresolved_required_fields"]:
        errors.append("必須項目の差し込み先を解決できませんでした: " + ", ".join(report["unresolved_required_fields"]))
    if errors:
        raise MappingResolutionError(" / ".join(errors))


def summarize_resolution_report(report: dict) -> str:
    return (
        f"resolved={len(report['resolved_fields'])}, "
        f"missing_required_data={len(report['missing_required_data'])}, "
        f"unresolved_required={len(report['unresolved_required_fields'])}, "
        f"unresolved_optional={len(report['unresolved_optional_fields'])}, "
        f"reference_skipped={len(report['reference_sheets_skipped'])}"
    )
