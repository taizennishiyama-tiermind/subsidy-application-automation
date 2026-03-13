"""Excel テンプレート処理の共通ユーティリティ

【書き込み方式】
openpyxl は DrawingML・DataValidation 等を削除してしまい、
複雑な公式テンプレートが破損・開けなくなる問題がある。

そのため fill_xlsx_safe() では ZIP 直接編集方式を採用:
  1. テンプレートを出力先にコピー
  2. ZIP を展開し sharedStrings.xml にテキストを追加
  3. sheet XML のセル値のみ書き換え（t="s" で sharedStrings を参照）
  4. 他の全コンポーネント（図形・入力規則等）はそのまま保持
  → 元ファイルと同等の xlsx が生成され、Excel/Numbers で正常に開ける

【名前空間プレフィックス保持】
Python の ElementTree は XML を再シリアライズ時に名前空間プレフィックスを
独自の ns0/ns1/ns2... に変更してしまう（例: mc: → ns1:, x14ac: → ns3:）。
このため xlsx の sheet XML が破損し Excel が読み込めなくなる。
→ セルの書き換えには ET を使わず、正規表現ベースの文字列操作を使用。

【sharedStrings リッチテキスト保持】
ET でリビルドすると <r><rPr>...<t>text</t></r> 形式のリッチテキストが
平文 <t></t> に潰される（または空文字列になる）。
→ 元の sharedStrings.xml バイト列を保持し、新文字列のみ末尾に追記する。
"""

import re
import sys
import zipfile
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import coordinate_to_tuple, column_index_from_string

_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
ET.register_namespace("", _NS)


# ---------------------------------------------------------------------------
# 内部ユーティリティ
# ---------------------------------------------------------------------------

def _parse_cell_ref(cell_ref: str) -> tuple[int, int]:
    """'K10' → (row=10, col=11)"""
    m = re.match(r"([A-Za-z]+)(\d+)", cell_ref.upper())
    if not m:
        raise ValueError(f"不正なセル参照: {cell_ref}")
    return int(m.group(2)), column_index_from_string(m.group(1))


def _get_or_add_shared_string(strings: list[str], text: str) -> int:
    """テキストのインデックスを返す。なければ追加する"""
    try:
        return strings.index(text)
    except ValueError:
        strings.append(text)
        return len(strings) - 1


def _extract_si_text(si_elem: ET.Element) -> str:
    """
    <si> 要素からプレーンテキストを抽出する。
    シンプル文字列 (<si><t>text</t></si>) と
    リッチテキスト (<si><r><rPr>...</rPr><t>text</t></r>...</si>) の両方に対応。
    """
    # シンプル文字列: <si><t>text</t></si>
    t = si_elem.find(f"{{{_NS}}}t")
    if t is not None:
        return t.text or ""
    # リッチテキスト: <si><r>...<t>text</t></r>...</si>
    parts = []
    for r in si_elem.findall(f"{{{_NS}}}r"):
        t = r.find(f"{{{_NS}}}t")
        if t is not None and t.text:
            parts.append(t.text)
    return "".join(parts)


def _get_sheet_xml_path(zf: zipfile.ZipFile, sheet_name: str) -> str | None:
    """シート名 → ZIP 内の XML パスを返す"""
    RNS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
    sheet_id = None
    for s in wb_xml.findall(".//{%s}sheet" % _NS):
        if s.get("name") == sheet_name:
            sheet_id = s.get("{%s}id" % RNS)
            break
    if sheet_id is None:
        return None
    rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    for rel in rels_xml:
        if rel.get("Id") == sheet_id:
            target = rel.get("Target")
            return target if target.startswith("xl/") else "xl/" + target
    return None


def _load_shared_strings(zf: zipfile.ZipFile) -> tuple[list[str], bytes]:
    """
    sharedStrings.xml を読み込み (テキストリスト, 元のXMLバイト列) を返す。
    テキストリストはインデックス検索用。元のXMLバイト列は保全用（再構築しない）。
    """
    path = "xl/sharedStrings.xml"
    if path not in zf.namelist():
        empty_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            ' count="0" uniqueCount="0"></sst>'
        ).encode("utf-8")
        return [], empty_xml
    raw = zf.read(path)
    root = ET.fromstring(raw)
    strings = [_extract_si_text(si) for si in root.findall(f"{{{_NS}}}si")]
    return strings, raw


def _build_shared_strings_xml(
    strings: list[str], original_raw: bytes, original_count: int
) -> bytes:
    """
    元の sharedStrings.xml バイト列を保持したまま、新規文字列のみ末尾に追記する。
    リッチテキスト・書式設定は一切変更しない。
    """
    new_strings = strings[original_count:]
    total = len(strings)

    xml = original_raw.decode("utf-8")
    # count / uniqueCount を更新
    xml = re.sub(r'\bcount="\d+"', f'count="{total}"', xml)
    xml = re.sub(r'\buniqueCount="\d+"', f'uniqueCount="{total}"', xml)

    if not new_strings:
        return xml.encode("utf-8")

    # 新しい <si> 要素を </sst> 直前に追加（XML エスケープ済み）
    # 改行文字・前後スペースがある場合は xml:space="preserve" を付与する
    parts = []
    for s in new_strings:
        s_str = str(s)
        s_esc = (
            s_str.replace("&", "&amp;")
                 .replace("<", "&lt;")
                 .replace(">", "&gt;")
        )
        needs_preserve = (
            s_str and (
                s_str[0] == " "
                or s_str[-1] == " "
                or "\n" in s_str
            )
        )
        if needs_preserve:
            parts.append(f'<si><t xml:space="preserve">{s_esc}</t></si>')
        else:
            parts.append(f'<si><t>{s_esc}</t></si>')

    xml = xml.replace("</sst>", "".join(parts) + "</sst>")
    return xml.encode("utf-8")


def _write_cell_in_sheet_xml(
    xml: str,
    ref_upper: str,
    row_num: int,
    col_idx: int,
    v_text: str,
    type_attr: str | None,
) -> str:
    """
    シート XML 文字列内の指定セルに値を書き込む（文字列操作ベース）。
    ElementTree を経由しないため名前空間プレフィックスが変わらない。

    処理の流れ:
      1. 既存セルを正規表現で探して値を上書き
      2. セルが無ければ、行要素に列順を守って挿入
      3. 行も無ければ sheetData に行ごと挿入
    """
    # --- 1. 既存セルを探して上書き ---
    # <c r="REF" .../> または <c r="REF" ...>...</c> の両方にマッチ
    cell_open_re = re.compile(
        r"<c\b[^>]*\br=\"" + re.escape(ref_upper) + r"\"[^>]*(?:/>|>)",
        re.DOTALL,
    )
    m = cell_open_re.search(xml)

    if m:
        tag = m.group(0)
        start = m.start()

        if tag.rstrip().endswith("/>"):
            # 自己閉じタグ: コンテンツなし
            end = m.end()
        else:
            # 対応する </c> を探す（<c> は入れ子にならない）
            close_pos = xml.find("</c>", m.end())
            if close_pos == -1:
                return xml  # 不正な XML はスキップ
            end = close_pos + len("</c>")

        # 元タグから属性部分を抽出し、t= を除去してから再構築
        attrs_match = re.search(r"^<c\b(.*?)(?:/>|>)$", tag, re.DOTALL)
        if not attrs_match:
            return xml
        raw_attrs = attrs_match.group(1)
        raw_attrs = re.sub(r"\s*\bt=\"[^\"]*\"", "", raw_attrs)  # t= 除去

        if type_attr:
            new_cell = f'<c{raw_attrs} t="{type_attr}"><v>{v_text}</v></c>'
        else:
            new_cell = f'<c{raw_attrs}><v>{v_text}</v></c>'

        return xml[:start] + new_cell + xml[end:]

    # --- 2. セルが無い: 既存の行に列順で挿入 ---
    if type_attr:
        new_cell = f'<c r="{ref_upper}" t="{type_attr}"><v>{v_text}</v></c>'
    else:
        new_cell = f'<c r="{ref_upper}"><v>{v_text}</v></c>'

    row_re = re.compile(
        r"(<row\b[^>]*\br=\"" + str(row_num) + r"\"[^>]*>)(.*?)(</row>)",
        re.DOTALL,
    )
    rm = row_re.search(xml)
    if rm:
        row_content = rm.group(2)
        insert_pos = len(row_content)
        for cm in re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)"', row_content):
            try:
                _, c_col = _parse_cell_ref(cm.group(1))
                if c_col > col_idx:
                    insert_pos = cm.start()
                    break
            except ValueError:
                continue
        new_row = rm.group(1) + row_content[:insert_pos] + new_cell + row_content[insert_pos:] + rm.group(3)
        return xml[: rm.start()] + new_row + xml[rm.end() :]

    # --- 3. 行も無い: sheetData に行ごと挿入 ---
    sd_re = re.compile(r"(<sheetData[^>]*>)(.*?)(</sheetData>)", re.DOTALL)
    sdm = sd_re.search(xml)
    if not sdm:
        return xml

    new_row = f'<row r="{row_num}">{new_cell}</row>'
    sd_content = sdm.group(2)
    insert_pos = len(sd_content)
    for er in re.finditer(r'<row\b[^>]*\br="(\d+)"', sd_content):
        if int(er.group(1)) > row_num:
            insert_pos = er.start()
            break

    new_sd = sd_content[:insert_pos] + new_row + sd_content[insert_pos:]
    return xml[: sdm.start()] + sdm.group(1) + new_sd + sdm.group(3) + xml[sdm.end() :]


def _get_merged_top_left(zf: zipfile.ZipFile, sheet_xml_path: str) -> dict[str, str]:
    """
    sheet XML を解析し、結合セル範囲内の非左上セル → 左上セルのマッピングを返す。
    例: {"B2": "A1", "C2": "A1", "A2": "A1"}  (A1:C2 が結合されている場合)
    これにより、結合セルの非左上セルへの書き込みを自動的に左上セルにリダイレクトできる。
    """
    from openpyxl.utils import get_column_letter

    raw = zf.read(sheet_xml_path)
    root = ET.fromstring(raw)
    merged_map: dict[str, str] = {}

    for mc in root.findall(f".//{{{_NS}}}mergeCell"):
        ref = mc.get("ref", "")
        parts = ref.split(":")
        if len(parts) != 2:
            continue
        top_left = parts[0].upper()
        try:
            r1, c1 = _parse_cell_ref(top_left)
            r2, c2 = _parse_cell_ref(parts[1].upper())
        except ValueError:
            continue
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell_ref = f"{get_column_letter(c)}{r}"
                if cell_ref != top_left:
                    merged_map[cell_ref] = top_left

    return merged_map


def _write_cells_to_sheet_xml(
    xml_bytes: bytes,
    cell_values: dict,
    strings: list[str],
    merged_map: dict[str, str] | None = None,
) -> tuple[bytes, list[str], list[str]]:
    """
    sheet XML バイト列の指定セルに値を書き込んで返す。
    文字列操作ベースで名前空間プレフィックスを変更しない。
    数値は t 属性なし <v> に直接書き込み、文字列は sharedStrings 経由 (t="s")。

    戻り値: (更新後のXMLバイト列, 書き込み成功リスト, リダイレクトログリスト)

    改善点:
    - merged_map があれば結合セルの非左上セルへの書き込みを自動的に左上セルへリダイレクト
    - 文字列内の改行 (\\n) を xml:space="preserve" として正しく処理
    """
    xml = xml_bytes.decode("utf-8")
    written: list[str] = []
    redirected: list[str] = []

    for ref, val in cell_values.items():
        ref_upper = ref.upper()

        # 結合セルの非左上セルへの書き込みをリダイレクト
        if merged_map and ref_upper in merged_map:
            original = ref_upper
            ref_upper = merged_map[ref_upper]
            redirected.append(f"{original} → {ref_upper} (結合セル自動リダイレクト)")

        row_num, col_idx = _parse_cell_ref(ref_upper)

        if isinstance(val, (int, float)):
            v_text = str(val)
            type_attr = None
        else:
            idx = _get_or_add_shared_string(strings, str(val))
            v_text = str(idx)
            type_attr = "s"

        xml = _write_cell_in_sheet_xml(xml, ref_upper, row_num, col_idx, v_text, type_attr)
        written.append(ref_upper)

    return xml.encode("utf-8"), written, redirected


# ---------------------------------------------------------------------------
# メイン書き込み関数（外部から呼び出す）
# ---------------------------------------------------------------------------

def fill_xlsx_safe(
    template_path: str,
    sheet_cell_map: dict,
    output_path: str,
    verbose: bool = True,
) -> dict:
    """
    テンプレートの構造を完全に保持したまま、指定セルに値を書き込む。

    sheet_cell_map: {
        "シート名": {"K10": "値", "AA23": 8000, ...},
        ...
    }
    戻り値: {"written": [...], "skipped": [...], "redirected": [...]}

    改善点（精度向上）:
    - 結合セルの非左上セルへの書き込みを自動検出して左上セルへリダイレクト
    - 文字列内の改行 (\\n) を xml:space="preserve" として正しく保存
    - 書き込み結果の検証レポートを出力
    - 名前空間プレフィックスを変更しない文字列操作方式を維持
    - sharedStrings のリッチテキスト・書式を保持（元XML を再利用）
    - calcChain.xml を削除してExcelが正しく再計算するよう促す
    """
    src = Path(template_path)
    dst = Path(output_path)
    dst.parent.mkdir(parents=True, exist_ok=True)

    report: dict = {"written": [], "skipped": [], "redirected": []}

    with zipfile.ZipFile(src, "r") as zf_src:
        # シート名 → XML パスを解決
        sheet_paths = {}
        for sheet_name in sheet_cell_map:
            path = _get_sheet_xml_path(zf_src, sheet_name)
            if path is None:
                print(f"⚠️ 警告: シート '{sheet_name}' が見つかりません。スキップします。")
                report["skipped"].append(f"シート '{sheet_name}' 未検出")
            else:
                sheet_paths[sheet_name] = path

        # 結合セルマップを各シートで取得
        merged_maps = {}
        for sheet_name, xml_path in sheet_paths.items():
            merged_maps[sheet_name] = _get_merged_top_left(zf_src, xml_path)

        # sharedStrings を読み込み（元バイト列も保持する）
        strings, original_ss_raw = _load_shared_strings(zf_src)
        original_count = len(strings)

        # 全ファイルを読み込み（全シートを含む）
        all_files = {name: zf_src.read(name) for name in zf_src.namelist()}
        infos = {info.filename: info for info in zf_src.infolist()}

    # 対象シートの XML のみ書き換え（他シートは無変更）
    for sheet_name, cell_values in sheet_cell_map.items():
        if sheet_name not in sheet_paths:
            continue
        xml_path = sheet_paths[sheet_name]
        if xml_path not in all_files:
            print(f"⚠️ 警告: {xml_path} が ZIP 内に見つかりません。")
            continue
        merged_map = merged_maps.get(sheet_name, {})
        new_xml, written, redirected = _write_cells_to_sheet_xml(
            all_files[xml_path], cell_values, strings, merged_map
        )
        all_files[xml_path] = new_xml
        report["written"].extend([f"{sheet_name}!{ref}" for ref in written])
        report["redirected"].extend([f"{sheet_name}: {r}" for r in redirected])

    # sharedStrings.xml を更新（元の内容を保持して新規文字列のみ追記）
    all_files["xl/sharedStrings.xml"] = _build_shared_strings_xml(
        strings, original_ss_raw, original_count
    )

    # workbook.xml.rels に sharedStrings の参照がなければ追加
    rels_path = "xl/_rels/workbook.xml.rels"
    rels_xml = all_files[rels_path].decode("utf-8")
    if "sharedStrings" not in rels_xml:
        rels_xml = rels_xml.replace(
            "</Relationships>",
            '<Relationship Id="rIdSS" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
            'Target="sharedStrings.xml"/></Relationships>',
        )
        all_files[rels_path] = rels_xml.encode("utf-8")

    # [Content_Types].xml に sharedStrings エントリがなければ追加
    ct_path = "[Content_Types].xml"
    ct_xml = all_files[ct_path].decode("utf-8")
    if "sharedStrings" not in ct_xml:
        ct_xml = ct_xml.replace(
            "</Types>",
            '<Override PartName="/xl/sharedStrings.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
            "</Types>",
        )

    # calcChain.xml を削除（修正後の値で数式キャッシュがずれる問題を回避）
    # Excel は calcChain がない場合に初回起動時に再計算するため正しい結果になる
    all_files.pop("xl/calcChain.xml", None)
    ct_xml = re.sub(
        r'<Override\b[^>]*PartName="/xl/calcChain\.xml"[^>]*/>', "", ct_xml
    )
    all_files[ct_path] = ct_xml.encode("utf-8")

    # ZIP を再構築（全ファイルを元の圧縮設定で書き込む）
    with zipfile.ZipFile(dst, "w") as zf_dst:
        for name, data in all_files.items():
            info = infos.get(name)
            if info:
                zf_dst.writestr(info, data)
            else:
                zf_dst.writestr(name, data)

    if verbose:
        print(f"\n【書き込み検証レポート】")
        print(f"  ✅ 書き込み完了: {len(report['written'])} セル")
        for ref in report["written"]:
            print(f"     {ref}")
        if report["redirected"]:
            print(f"  ↩️  結合セル自動リダイレクト: {len(report['redirected'])} 件")
            for r in report["redirected"]:
                print(f"     {r}")
        if report["skipped"]:
            print(f"  ⚠️ スキップ: {len(report['skipped'])} 件")
            for s in report["skipped"]:
                print(f"     {s}")
        print(f"\n✅ 保存完了: {output_path}")
    else:
        print(f"保存完了: {output_path}")

    return report


# ---------------------------------------------------------------------------
# 構造解析（/add-template, analyze コマンド用）
# ---------------------------------------------------------------------------

def load_template(template_path: str):
    path = Path(template_path)
    if not path.exists():
        print(f"エラー: テンプレートファイルが見つかりません: {template_path}", file=sys.stderr)
        sys.exit(1)
    return load_workbook(template_path)


def _resolve_merged_cell(ws, cell_ref: str):
    cell = ws[cell_ref]
    if not isinstance(cell, MergedCell):
        return cell
    target_row, target_col = coordinate_to_tuple(cell_ref)
    for merge_range in ws.merged_cells.ranges:
        if (merge_range.min_row <= target_row <= merge_range.max_row and
                merge_range.min_col <= target_col <= merge_range.max_col):
            return ws.cell(row=merge_range.min_row, column=merge_range.min_col)
    return cell


def fill_cell(ws, cell_ref: str, value, preserve_format: bool = True):
    cell = _resolve_merged_cell(ws, cell_ref)
    cell.value = value


def fill_cells(ws, mappings: dict):
    for cell_ref, value in mappings.items():
        if value is not None and value != "":
            fill_cell(ws, cell_ref, value)


def save_output(wb, output_path: str):
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"保存完了: {output_path}")


def read_cell(ws, cell_ref: str) -> str:
    value = ws[cell_ref].value
    return str(value) if value is not None else ""


def read_cells(ws, cell_refs: list) -> dict:
    return {ref: read_cell(ws, ref) for ref in cell_refs}


def list_sheets(wb) -> list:
    return wb.sheetnames


def get_template_info(template_path: str) -> dict:
    wb = load_template(template_path)
    info = {}
    for name in wb.sheetnames:
        ws = wb[name]
        info[name] = {
            "min_row": ws.min_row,
            "max_row": ws.max_row,
            "min_col": ws.min_column,
            "max_col": ws.max_column,
            "dimensions": ws.dimensions,
        }
    return info
