#!/usr/bin/env python3
"""
Excel Template Fill Script
マッピング定義JSONに従い、テンプレートxlsxにデータを代入して完成版を出力する。

対応する代入パターン:
  1. Named Range への値代入
  2. セル参照（例: Sheet1!B2）への直接代入
  3. 表への動的行追加（行挿入 + 書式コピー）

Usage:
  python fill_template.py \
    --template template.xlsx \
    --mapping  mapping.json \
    --values   values.json \
    --output   output.xlsx
"""

import argparse
import json
import os
import re
import copy
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


# ---------------------------------------------------------------------------
# 値のパース
# ---------------------------------------------------------------------------

def parse_value(raw: Any, data_type: str = "auto") -> Any:
    """値を適切なPython型に変換する"""
    if raw is None:
        return None

    if data_type == "string":
        return str(raw)
    if data_type == "number":
        try:
            return int(raw) if '.' not in str(raw) else float(raw)
        except (ValueError, TypeError):
            return raw
    if data_type == "date":
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日']:
            try:
                return datetime.strptime(str(raw), fmt)
            except ValueError:
                continue
        return str(raw)
    if data_type == "formula":
        return raw  # 数式はそのまま

    # auto: 自動判定
    try:
        if '.' in str(raw):
            return float(raw)
        return int(raw)
    except (ValueError, TypeError):
        pass
    for fmt in ['%Y-%m-%d', '%Y/%m/%d']:
        try:
            return datetime.strptime(str(raw), fmt)
        except ValueError:
            continue
    return str(raw)


# ---------------------------------------------------------------------------
# セル参照のパース
# ---------------------------------------------------------------------------

def parse_cell_ref(ref: str):
    """'Sheet1!B2' → (sheet_name, cell_ref)  /  'B2' → (None, 'B2')"""
    if '!' in ref:
        sheet, cell = ref.split('!', 1)
        return sheet, cell.replace('$', '')
    return None, ref.replace('$', '')


# ---------------------------------------------------------------------------
# 書式コピー
# ---------------------------------------------------------------------------

def copy_cell_style(src_cell, dst_cell):
    """セルの書式をコピーする（値はコピーしない）"""
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy.copy(src_cell.protection)
        dst_cell.alignment = copy.copy(src_cell.alignment)


# ---------------------------------------------------------------------------
# Named Range の解決
# ---------------------------------------------------------------------------

def resolve_named_range(wb, name: str):
    """Named Range名 → (sheet_name, cell_ref) を返す。見つからなければ None"""
    if name not in wb.defined_names:
        return None
    dn = wb.defined_names[name]
    dests = list(dn.destinations)
    if not dests:
        return None
    sheet_name, cell_ref = dests[0]
    return sheet_name, cell_ref.replace('$', '')


# ---------------------------------------------------------------------------
# 単一セル代入
# ---------------------------------------------------------------------------

def fill_cells(wb, cells_def: list, values: dict) -> tuple[list, list]:
    """cells 定義に従って単一セルに値を代入する"""
    filled = []
    skipped = []

    for cell_def in cells_def:
        key = cell_def['key']
        if key not in values:
            skipped.append(f"{key}: 値データに存在しません")
            continue

        raw_value = values[key]
        data_type = cell_def.get('data_type', 'auto')
        parsed = parse_value(raw_value, data_type)

        target_type = cell_def.get('target_type', 'named_range')
        target = cell_def['target']

        if target_type == 'named_range':
            result = resolve_named_range(wb, target)
            if result is None:
                skipped.append(f"{key}: Named Range '{target}' が見つかりません")
                continue
            sheet_name, cell_ref = result
        else:  # cell
            sheet_name, cell_ref = parse_cell_ref(target)
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]

        if sheet_name not in wb.sheetnames:
            skipped.append(f"{key}: シート '{sheet_name}' が存在しません")
            continue

        ws = wb[sheet_name]
        ws[cell_ref] = parsed
        label = cell_def.get('label', key)
        filled.append(f"{label} → {sheet_name}!{cell_ref} = {parsed}")

    return filled, skipped


# ---------------------------------------------------------------------------
# 表（動的行）代入
# ---------------------------------------------------------------------------

def fill_tables(wb, tables_def: list, values: dict) -> tuple[list, list]:
    """tables 定義に従って表に行を挿入・代入する"""
    filled = []
    skipped = []

    for table_def in tables_def:
        key = table_def['key']
        if key not in values:
            skipped.append(f"{key}: 値データに存在しません")
            continue

        rows_data = values[key]
        if not isinstance(rows_data, list):
            skipped.append(f"{key}: 配列が必要です")
            continue

        sheet_name = table_def['sheet']
        if sheet_name not in wb.sheetnames:
            skipped.append(f"{key}: シート '{sheet_name}' が存在しません")
            continue

        ws = wb[sheet_name]
        start_row = table_def['start_row']
        columns = table_def['columns']
        format_source_row = table_def.get('format_source_row', start_row)
        insert_mode = table_def.get('insert_mode', 'shift_down')
        num_rows = len(rows_data)

        # 書式コピー元の情報を事前取得
        format_cells = {}
        for col_def in columns:
            col_letter = col_def['col']
            src_cell = ws[f"{col_letter}{format_source_row}"]
            format_cells[col_letter] = src_cell

        # 行挿入（shift_down の場合、2行目以降の分だけ挿入）
        if insert_mode == 'shift_down' and num_rows > 1:
            ws.insert_rows(start_row + 1, amount=num_rows - 1)

        # データ書き込み
        table_log = []
        for i, row_data in enumerate(rows_data):
            current_row = start_row + i
            row_values = []

            for col_def in columns:
                col_letter = col_def['col']
                col_key = col_def['key']
                data_type = col_def.get('data_type', 'auto')
                cell_ref = f"{col_letter}{current_row}"

                # 書式コピー
                src = format_cells[col_letter]
                dst = ws[cell_ref]
                copy_cell_style(src, dst)

                # 値の書き込み
                if data_type == 'formula':
                    formula_tmpl = col_def.get('formula_template', '')
                    formula = formula_tmpl.replace('{row}', str(current_row))
                    ws[cell_ref] = formula
                    row_values.append(formula)
                elif col_key in row_data:
                    parsed = parse_value(row_data[col_key], data_type)
                    ws[cell_ref] = parsed
                    row_values.append(str(parsed))
                else:
                    row_values.append('')

            table_log.append(f"  行{current_row}: {' | '.join(row_values)}")

        # 合計行のSUM範囲を自動調整
        if insert_mode == 'shift_down' and num_rows > 1:
            _adjust_sum_formulas(ws, start_row, num_rows, columns)

        label = table_def.get('label', key)
        filled.append(f"表「{label}」挿入完了 ({num_rows}行):")
        filled.extend(table_log)

    return filled, skipped


def _adjust_sum_formulas(ws, start_row: int, num_rows: int, columns: list):
    """
    表の下にあるSUM数式の範囲を調整する。
    挿入した行数分だけ検索範囲を広げる。
    """
    end_data_row = start_row + num_rows - 1
    search_range = range(end_data_row + 1, min(end_data_row + 10, ws.max_row + 1))

    for row_num in search_range:
        for col_def in columns:
            col_letter = col_def['col']
            cell = ws[f"{col_letter}{row_num}"]
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                # SUM(X10:X10) のようなパターンを SUM(X10:X{end}) に修正
                pattern = rf'(SUM\({col_letter}){start_row}(:{col_letter}){start_row}(\))'
                replacement = rf'\g<1>{start_row}\g<2>{end_data_row}\3'
                new_formula = re.sub(pattern, replacement, formula, flags=re.IGNORECASE)
                if new_formula != formula:
                    cell.value = new_formula


# ---------------------------------------------------------------------------
# メイン処理
# ---------------------------------------------------------------------------

def fill_template(template_path: str, mapping_path: str, values_path: str, output_path: str):
    """テンプレートにデータを代入して出力する"""

    # ファイル読み込み
    with open(mapping_path, 'r', encoding='utf-8') as f:
        mapping = json.load(f)
    with open(values_path, 'r', encoding='utf-8') as f:
        values = json.load(f)

    wb = openpyxl.load_workbook(template_path)

    all_filled = []
    all_skipped = []

    # 1) 単一セル代入
    if 'cells' in mapping:
        filled, skipped = fill_cells(wb, mapping['cells'], values)
        all_filled.extend(filled)
        all_skipped.extend(skipped)

    # 2) 表の代入
    if 'tables' in mapping:
        filled, skipped = fill_tables(wb, mapping['tables'], values)
        all_filled.extend(filled)
        all_skipped.extend(skipped)

    # 出力
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    wb.save(output_path)

    # ログ出力
    print(f"\n✅ 代入完了 ({len(all_filled)}件):")
    for line in all_filled:
        print(f"  {line}")

    if all_skipped:
        print(f"\n⚠️ スキップ ({len(all_skipped)}件):")
        for line in all_skipped:
            print(f"  {line}")
    else:
        print(f"\n⚠️ スキップ (0件)")

    print(f"\n📄 出力: {output_path}")
    return all_filled, all_skipped


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Excel Template Fill')
    parser.add_argument('--template', required=True, help='テンプレートxlsxのパス')
    parser.add_argument('--mapping',  required=True, help='マッピング定義JSONのパス')
    parser.add_argument('--values',   required=True, help='値データJSONのパス')
    parser.add_argument('--output',   required=True, help='出力先xlsxのパス')
    args = parser.parse_args()

    fill_template(args.template, args.mapping, args.values, args.output)
